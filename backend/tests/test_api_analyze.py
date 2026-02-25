"""Tests for POST /api/analyze — file upload, pipeline, xlsx response."""

import io
import json
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.main import app
from app.models.database import Base, get_db
from app.models.seed import seed_database

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

_FIXTURES = Path(__file__).parent / "fixtures"
_RVTOOLS = _FIXTURES / "sample_rvtools.xlsx"
_CLOUDPHYSICS = _FIXTURES / "sample_cloudphysics.xlsx"


@pytest.fixture
def client():
    """TestClient with a fresh in-memory seeded DB injected via dependency override."""
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    db = Session()
    seed_database(db)

    def override_get_db():
        yield db

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.pop(get_db, None)
    db.close()
    Base.metadata.drop_all(engine)


def _upload(client: TestClient, path: Path, customer_name: str = None):
    """Helper to POST a file to /api/analyze."""
    with open(path, "rb") as f:
        data = {}
        if customer_name:
            data["customer_name"] = customer_name
        return client.post(
            "/api/analyze",
            files={
                "file": (
                    path.name,
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data=data,
        )


# ---------------------------------------------------------------------------
# Success cases
# ---------------------------------------------------------------------------


def test_analyze_rvtools_returns_xlsx(client: TestClient) -> None:
    resp = _upload(client, _RVTOOLS)
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    wb = load_workbook(io.BytesIO(resp.content))
    assert "VM Detail" in wb.sheetnames


def test_analyze_cloudphysics_returns_xlsx(client: TestClient) -> None:
    resp = _upload(client, _CLOUDPHYSICS)
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    assert "VM Detail" in wb.sheetnames


def test_analyze_returns_analysis_summary_header(client: TestClient) -> None:
    resp = _upload(client, _RVTOOLS)
    assert resp.status_code == 200
    assert "x-analysis-summary" in resp.headers
    summary = json.loads(resp.headers["x-analysis-summary"])
    assert "total" in summary
    for tier in [
        "officially_supported",
        "unofficially_supported",
        "supported_vdi",
        "needs_review",
        "needs_info",
        "not_supported",
    ]:
        assert tier in summary
    tier_sum = sum(
        summary[t]
        for t in [
            "officially_supported",
            "unofficially_supported",
            "supported_vdi",
            "needs_review",
            "needs_info",
            "not_supported",
        ]
    )
    assert summary["total"] == tier_sum


def test_analyze_filename_includes_customer_name(client: TestClient) -> None:
    resp = _upload(client, _RVTOOLS, customer_name="Acme Corp")
    assert resp.status_code == 200
    disposition = resp.headers.get("content-disposition", "")
    assert "Acme-Corp" in disposition
    assert ".xlsx" in disposition


def test_analyze_filename_without_customer_name(client: TestClient) -> None:
    resp = _upload(client, _RVTOOLS)
    assert resp.status_code == 200
    disposition = resp.headers.get("content-disposition", "")
    assert "VME-Analysis-" in disposition
    assert "None" not in disposition


# ---------------------------------------------------------------------------
# Error cases
# ---------------------------------------------------------------------------


def test_analyze_rejects_oversized_file(client: TestClient) -> None:
    big_bytes = b"x" * (11 * 1024 * 1024)
    resp = client.post(
        "/api/analyze",
        files={"file": ("big.xlsx", io.BytesIO(big_bytes), "application/octet-stream")},
    )
    assert resp.status_code == 413
    body = resp.json()
    assert body["error"] is True
    assert body["code"] == "FILE_TOO_LARGE"


def test_analyze_rejects_non_xlsx(client: TestClient) -> None:
    csv_bytes = b"VM Name,Guest OS\nmy-vm,Windows Server 2022\n"
    resp = client.post(
        "/api/analyze",
        files={"file": ("inventory.csv", io.BytesIO(csv_bytes), "text/csv")},
    )
    assert resp.status_code == 400
    body = resp.json()
    assert body["code"] == "UNRECOGNIZED_FORMAT"


def test_analyze_rejects_unrecognized_xlsx(client: TestClient) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Foo", "Bar", "Baz"])
    ws.append(["a", "b", "c"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = client.post(
        "/api/analyze",
        files={
            "file": (
                "unknown.xlsx",
                buf,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp.status_code == 400
    assert resp.json()["code"] == "UNRECOGNIZED_FORMAT"


def test_analyze_missing_columns(client: TestClient) -> None:
    # Passes format detection (has all 3 signature cols) but missing
    # "OS according to configuration file" (required for extraction)
    wb = Workbook()
    ws = wb.active
    ws.title = "vInfo"
    ws.append(["VM", "Powerstate", "OS according to VMware Tools"])
    ws.append(["my-vm", "poweredOn", "Windows Server 2022"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = client.post(
        "/api/analyze",
        files={
            "file": (
                "partial.xlsx",
                buf,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp.status_code == 400
    body = resp.json()
    assert body["code"] == "MISSING_COLUMNS"
    assert "missing" in body


# ---------------------------------------------------------------------------
# Error response shape
# ---------------------------------------------------------------------------


def test_error_response_has_error_code_and_message(client: TestClient) -> None:
    resp = client.post(
        "/api/analyze",
        files={"file": ("bad.csv", io.BytesIO(b"garbage"), "text/csv")},
    )
    assert resp.status_code == 400
    body = resp.json()
    assert body["error"] is True
    assert isinstance(body["code"], str)
    assert isinstance(body["message"], str)
