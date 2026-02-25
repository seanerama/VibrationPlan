"""Tests for OutputBuilder — xlsx generation, formatting, and filename logic."""

import io
from typing import Optional

import pytest
from openpyxl import load_workbook

from app.models.constants import (
    ALL_TIERS,
    TIER_DISPLAY_NAMES,
    TIER_OFFICIALLY_SUPPORTED,
    TIER_ROW_BG,
)
from app.services.classification_engine import ClassifiedVM
from app.services.output_builder import OutputBuilder

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _vm(
    vm_name: str = "test-vm",
    tier: str = TIER_OFFICIALLY_SUPPORTED,
    os_raw: str = "Windows Server 2022",
    os_interpreted: str = "Windows Server 2022",
    host_cluster: str = "Cluster-A",
    reason: str = "HPE validated",
    migration_path: str = "Proceed with migration.",
    notes: Optional[str] = None,
) -> ClassifiedVM:
    from app.models.constants import TIER_COLORS

    return ClassifiedVM(
        vm_name=vm_name,
        host_cluster=host_cluster,
        os_raw=os_raw,
        os_interpreted=os_interpreted,
        classification_tier=tier,
        classification_color=TIER_COLORS[tier],
        classification_reason=reason,
        migration_path=migration_path,
        notes=notes,
    )


def _all_tier_vms() -> list[ClassifiedVM]:
    """One VM per tier — useful for tests that need all 6 tiers present."""
    return [_vm(vm_name=f"vm-{t}", tier=t) for t in ALL_TIERS]


@pytest.fixture
def builder() -> OutputBuilder:
    return OutputBuilder()


@pytest.fixture
def simple_vms() -> list[ClassifiedVM]:
    return [
        _vm("web-01", TIER_OFFICIALLY_SUPPORTED),
        _vm("db-01", "unofficially_supported"),
        _vm("web-01b", TIER_OFFICIALLY_SUPPORTED),
    ]


@pytest.fixture
def simple_bytes(builder: OutputBuilder, simple_vms: list[ClassifiedVM]) -> bytes:
    return builder.build(simple_vms)


@pytest.fixture
def simple_wb(simple_bytes: bytes):
    return load_workbook(io.BytesIO(simple_bytes))


# ---------------------------------------------------------------------------
# Basic output
# ---------------------------------------------------------------------------


def test_build_returns_bytes(builder: OutputBuilder, simple_vms: list[ClassifiedVM]) -> None:
    result = builder.build(simple_vms)
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_output_is_valid_xlsx(simple_bytes: bytes) -> None:
    wb = load_workbook(io.BytesIO(simple_bytes))
    assert wb is not None


def test_output_has_three_tabs(simple_wb) -> None:
    assert set(simple_wb.sheetnames) == {"VM Detail", "Summary", "Executive Summary"}


# ---------------------------------------------------------------------------
# VM Detail tab
# ---------------------------------------------------------------------------


def test_vm_detail_column_headers_correct(simple_wb) -> None:
    ws = simple_wb["VM Detail"]
    expected = [
        "VM Name", "Host/Cluster", "OS (Raw)", "OS (Interpreted)",
        "Classification", "Classification Color", "Classification Reason",
        "Migration Path", "Notes",
    ]
    # Header is on row 2 (row 1 = logo)
    actual = [ws.cell(row=2, column=i).value for i in range(1, len(expected) + 1)]
    assert actual == expected


def test_vm_detail_row_count_matches_input(
    builder: OutputBuilder, simple_vms: list[ClassifiedVM]
) -> None:
    wb = load_workbook(io.BytesIO(builder.build(simple_vms)))
    ws = wb["VM Detail"]
    # max_row includes logo row (1) + header row (1) + data rows
    data_rows = ws.max_row - 2
    assert data_rows == len(simple_vms)


def test_row_fill_matches_tier_color(builder: OutputBuilder) -> None:
    vms = [_vm(tier=TIER_OFFICIALLY_SUPPORTED)]
    wb = load_workbook(io.BytesIO(builder.build(vms)))
    ws = wb["VM Detail"]
    # First data row is row 3
    data_cell = ws.cell(row=3, column=1)
    assert data_cell.fill.fgColor.rgb == TIER_ROW_BG[TIER_OFFICIALLY_SUPPORTED]


def test_classification_column_uses_display_name(builder: OutputBuilder) -> None:
    vms = [_vm(tier=TIER_OFFICIALLY_SUPPORTED)]
    wb = load_workbook(io.BytesIO(builder.build(vms)))
    ws = wb["VM Detail"]
    # Classification is column 5, data starts at row 3
    class_value = ws.cell(row=3, column=5).value
    assert class_value == "Officially Supported"
    assert class_value != "officially_supported"


def test_vm_name_written_correctly(builder: OutputBuilder) -> None:
    vms = [_vm(vm_name="my-special-vm")]
    wb = load_workbook(io.BytesIO(builder.build(vms)))
    ws = wb["VM Detail"]
    assert ws.cell(row=3, column=1).value == "my-special-vm"


def test_freeze_pane_set_on_vm_detail(simple_wb) -> None:
    ws = simple_wb["VM Detail"]
    # Freeze pane should be at A3 (below header row 2)
    assert ws.freeze_panes is not None
    assert str(ws.freeze_panes) == "A3"


def test_notes_written_when_present(builder: OutputBuilder) -> None:
    vms = [_vm(notes="Low confidence match: 0.65")]
    wb = load_workbook(io.BytesIO(builder.build(vms)))
    ws = wb["VM Detail"]
    notes_cell = ws.cell(row=3, column=9)
    assert notes_cell.value == "Low confidence match: 0.65"


def test_notes_empty_string_when_none(builder: OutputBuilder) -> None:
    vms = [_vm(notes=None)]
    wb = load_workbook(io.BytesIO(builder.build(vms)))
    ws = wb["VM Detail"]
    notes_cell = ws.cell(row=3, column=9)
    assert notes_cell.value in ("", None)


# ---------------------------------------------------------------------------
# Summary tab
# ---------------------------------------------------------------------------


def test_summary_tab_has_all_six_tiers(builder: OutputBuilder) -> None:
    vms = _all_tier_vms()
    wb = load_workbook(io.BytesIO(builder.build(vms)))
    ws = wb["Summary"]
    # Tier names appear in column A starting at row 3
    tier_names_in_sheet = [ws.cell(row=r, column=1).value for r in range(3, 9)]
    for display_name in TIER_DISPLAY_NAMES.values():
        assert display_name in tier_names_in_sheet


def test_summary_counts_are_correct(builder: OutputBuilder) -> None:
    vms = [
        _vm(tier=TIER_OFFICIALLY_SUPPORTED),
        _vm(tier=TIER_OFFICIALLY_SUPPORTED),
        _vm(tier="not_supported"),
    ]
    wb = load_workbook(io.BytesIO(builder.build(vms)))
    ws = wb["Summary"]
    counts = {
        ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
        for r in range(3, 9)
    }
    assert counts["Officially Supported"] == 2
    assert counts["Not Supported"] == 1


def test_summary_percentages_sum_to_100(builder: OutputBuilder) -> None:
    vms = _all_tier_vms()  # 1 VM per tier = 6 total, each ~16.7%
    wb = load_workbook(io.BytesIO(builder.build(vms)))
    ws = wb["Summary"]
    pct_values = []
    for r in range(3, 9):
        val = ws.cell(row=r, column=3).value  # e.g. "16.7%"
        if val:
            pct_values.append(float(str(val).rstrip("%")))
    total_pct = round(sum(pct_values), 0)
    assert total_pct == 100.0


def test_summary_totals_row_present(builder: OutputBuilder) -> None:
    vms = _all_tier_vms()
    wb = load_workbook(io.BytesIO(builder.build(vms)))
    ws = wb["Summary"]
    # Totals row is row 9 (2 header + 6 tiers + 1)
    assert ws.cell(row=9, column=1).value == "TOTAL"
    assert ws.cell(row=9, column=2).value == len(vms)


# ---------------------------------------------------------------------------
# Executive Summary tab
# ---------------------------------------------------------------------------


def test_executive_summary_contains_vm_count(builder: OutputBuilder) -> None:
    vms = _all_tier_vms()
    wb = load_workbook(io.BytesIO(builder.build(vms)))
    ws = wb["Executive Summary"]
    all_text = " ".join(
        str(ws.cell(row=r, column=1).value or "")
        for r in range(1, ws.max_row + 1)
    )
    assert str(len(vms)) in all_text


def test_executive_summary_contains_customer_name(builder: OutputBuilder) -> None:
    vms = _all_tier_vms()
    wb = load_workbook(io.BytesIO(builder.build(vms, customer_name="Acme Corp")))
    ws = wb["Executive Summary"]
    all_text = " ".join(
        str(ws.cell(row=r, column=1).value or "")
        for r in range(1, ws.max_row + 1)
    )
    assert "Acme Corp" in all_text


def test_executive_summary_no_customer_name(builder: OutputBuilder) -> None:
    vms = _all_tier_vms()
    result = builder.build(vms, customer_name=None)
    wb = load_workbook(io.BytesIO(result))
    assert "Executive Summary" in wb.sheetnames


# ---------------------------------------------------------------------------
# Filename generation
# ---------------------------------------------------------------------------


def test_filename_includes_customer_name() -> None:
    name = OutputBuilder.build_filename("Acme Corp")
    assert "Acme-Corp" in name
    assert name.endswith(".xlsx")


def test_filename_without_customer_name() -> None:
    name = OutputBuilder.build_filename()
    assert name.startswith("VME-Analysis-")
    assert name.endswith(".xlsx")
    assert "None" not in name


def test_filename_sanitizes_special_chars() -> None:
    name = OutputBuilder.build_filename("Acme & Corp! (Test)")
    # Special chars should be stripped, spaces → dashes
    assert "&" not in name
    assert "!" not in name
    assert "(" not in name
    assert name.endswith(".xlsx")


def test_filename_contains_date() -> None:
    from datetime import datetime

    today = datetime.now().strftime("%Y-%m-%d")
    name = OutputBuilder.build_filename("Test")
    assert today in name


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------


def test_build_with_empty_vm_list(builder: OutputBuilder) -> None:
    result = builder.build([])
    assert isinstance(result, bytes)
    wb = load_workbook(io.BytesIO(result))
    assert "VM Detail" in wb.sheetnames
    ws = wb["VM Detail"]
    # Only logo row + header row — no data rows
    assert ws.max_row <= 2


def test_build_with_all_six_tiers(builder: OutputBuilder) -> None:
    vms = _all_tier_vms()
    result = builder.build(vms)
    wb = load_workbook(io.BytesIO(result))
    ws = wb["VM Detail"]
    # 6 data rows + header + logo
    assert ws.max_row == len(vms) + 2


def test_all_tier_row_fills_applied(builder: OutputBuilder) -> None:
    """Every tier's row bg color must appear in the VM Detail sheet."""
    vms = _all_tier_vms()
    wb = load_workbook(io.BytesIO(builder.build(vms)))
    ws = wb["VM Detail"]
    fills_seen = {ws.cell(row=r, column=1).fill.fgColor.rgb for r in range(3, 3 + len(vms))}
    for tier in ALL_TIERS:
        assert TIER_ROW_BG[tier] in fills_seen
