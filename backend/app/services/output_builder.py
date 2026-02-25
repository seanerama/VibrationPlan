"""Output Builder — generates the branded 3-tab .xlsx report."""

import io
import logging
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.constants import (
    ALL_TIERS,
    TIER_DISPLAY_NAMES,
    TIER_NEEDS_INFO,
    TIER_NEEDS_REVIEW,
    TIER_NOT_SUPPORTED,
    TIER_OFFICIALLY_SUPPORTED,
    TIER_ROW_BG,
    TIER_SUPPORTED_VDI,
    TIER_UNOFFICIALLY_SUPPORTED,
)
from app.services.classification_engine import ClassifiedVM

logger = logging.getLogger(__name__)

_ASSETS_DIR = Path(__file__).parent.parent / "assets"
_LOGO_PATH = _ASSETS_DIR / "USSBlueBurst.png"

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

_HEADER_BG = "FF1F2937"
_HEADER_FONT_COLOR = "FFF9FAFB"
_LOGO_BG = "FF0A0D12"
_FONT_MONO = "IBM Plex Mono"
_FONT_SANS = "IBM Plex Sans"

_VM_DETAIL_HEADERS = [
    "VM Name",
    "Host/Cluster",
    "OS (Raw)",
    "OS (Interpreted)",
    "Classification",
    "Classification Color",
    "Classification Reason",
    "Migration Path",
    "Notes",
]

_COLUMN_WIDTHS = {
    "VM Name": 28,
    "Host/Cluster": 22,
    "OS (Raw)": 32,
    "OS (Interpreted)": 28,
    "Classification": 26,
    "Classification Color": 18,
    "Classification Reason": 38,
    "Migration Path": 34,
    "Notes": 28,
}

_SUMMARY_HEADERS = ["Classification", "Count", "Percentage"]

_EXEC_SUMMARY_TEMPLATE = (
    "VME Compatibility Analysis{customer_suffix}\n"
    "Analysis Date: {date}\n"
    "Total VMs Analyzed: {total}\n"
    "\n"
    "SUMMARY\n"
    "{officially_supported_count} VMs ({officially_supported_pct}%) are Officially Supported "
    "and validated by HPE for migration to VM Essentials.\n"
    "{unofficially_supported_count} VMs ({unofficially_supported_pct}%) are Unofficially "
    "Supported — KVM-compatible but not HPE-validated; recommend non-production testing.\n"
    "{supported_vdi_count} VMs ({supported_vdi_pct}%) are Supported VDI workloads "
    "(Citrix, Omnissa Horizon, or HP Anyware) validated for VME.\n"
    "{needs_review_count} VMs ({needs_review_pct}%) require review with the customer "
    "to confirm OS version or compatibility.\n"
    "{needs_info_count} VMs ({needs_info_pct}%) have insufficient OS data for "
    "classification — additional information required.\n"
    "{not_supported_count} VMs ({not_supported_pct}%) are Not Supported on the KVM "
    "hypervisor and will require OS upgrade, re-platforming, or retention on VMware.\n"
    "\n"
    "This analysis was generated using the HPE VM Essentials compatibility matrix. "
    "All classification results should be reviewed with the customer prior to migration planning."
)


# ---------------------------------------------------------------------------
# Public interface
# ---------------------------------------------------------------------------


class OutputBuilder:
    """Generates the branded 3-tab .xlsx report from a list of ClassifiedVM objects."""

    def build(
        self,
        classified_vms: list[ClassifiedVM],
        customer_name: Optional[str] = None,
    ) -> bytes:
        """Generate the branded xlsx report and return it as raw bytes.

        Args:
            classified_vms: Classified VM records from ClassificationEngine.
            customer_name: Optional customer name included in the report header.

        Returns:
            Raw bytes of the .xlsx file, ready to stream as an HTTP response.
        """
        wb = Workbook()

        self._build_vm_detail(wb, classified_vms)
        self._build_summary(wb, classified_vms)
        self._build_exec_summary(wb, classified_vms, customer_name)

        # Remove the default empty sheet created by Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def build_filename(customer_name: Optional[str] = None) -> str:
        """Return a safe filename for the generated report."""
        date_str = datetime.now().strftime("%Y-%m-%d")
        if customer_name:
            safe = re.sub(r"[^a-zA-Z0-9\s-]", "", customer_name).strip().replace(" ", "-")
            return f"VME-Analysis-{safe}-{date_str}.xlsx"
        return f"VME-Analysis-{date_str}.xlsx"

    # ------------------------------------------------------------------
    # Tab 1 — VM Detail
    # ------------------------------------------------------------------

    def _build_vm_detail(self, wb: Workbook, vms: list[ClassifiedVM]) -> None:
        ws = wb.create_sheet("VM Detail")

        # Logo row
        self._insert_logo(ws, row=1)

        # Header row (row 2)
        header_row = 2
        header_fill = PatternFill(
            start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid"
        )
        header_font = Font(bold=True, color=_HEADER_FONT_COLOR, name=_FONT_SANS)

        for col_idx, header in enumerate(_VM_DETAIL_HEADERS, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=False, vertical="center")
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = _COLUMN_WIDTHS.get(header, 20)

        # Freeze pane below header
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # Data rows (start at row 3)
        for row_offset, vm in enumerate(vms):
            row_num = header_row + 1 + row_offset
            row_fill = PatternFill(
                start_color=TIER_ROW_BG[vm.classification_tier],
                end_color=TIER_ROW_BG[vm.classification_tier],
                fill_type="solid",
            )
            data_font = Font(name=_FONT_MONO, color="FFF9FAFB")
            values = [
                vm.vm_name,
                vm.host_cluster or "",
                vm.os_raw,
                vm.os_interpreted,
                TIER_DISPLAY_NAMES[vm.classification_tier],
                vm.classification_color,
                vm.classification_reason,
                vm.migration_path,
                vm.notes or "",
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.fill = row_fill
                cell.font = data_font
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Row height for header
        ws.row_dimensions[header_row].height = 22

    # ------------------------------------------------------------------
    # Tab 2 — Summary
    # ------------------------------------------------------------------

    def _build_summary(self, wb: Workbook, vms: list[ClassifiedVM]) -> None:
        ws = wb.create_sheet("Summary")

        # Logo row
        self._insert_logo(ws, row=1)

        # Header row
        header_row = 2
        header_fill = PatternFill(
            start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid"
        )
        header_font = Font(bold=True, color=_HEADER_FONT_COLOR, name=_FONT_SANS)

        for col_idx, header in enumerate(_SUMMARY_HEADERS, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 12

        total = len(vms)
        counts = Counter(vm.classification_tier for vm in vms)

        for row_offset, tier in enumerate(ALL_TIERS):
            row_num = header_row + 1 + row_offset
            count = counts.get(tier, 0)
            pct = round(count / total * 100, 1) if total > 0 else 0.0

            row_fill = PatternFill(
                start_color=TIER_ROW_BG[tier],
                end_color=TIER_ROW_BG[tier],
                fill_type="solid",
            )
            row_font = Font(name=_FONT_SANS, color="FFF9FAFB")

            ws.cell(row=row_num, column=1, value=TIER_DISPLAY_NAMES[tier]).fill = row_fill
            ws.cell(row=row_num, column=2, value=count).fill = row_fill
            ws.cell(row=row_num, column=3, value=f"{pct}%").fill = row_fill

            for col_idx in range(1, 4):
                ws.cell(row=row_num, column=col_idx).font = row_font

        # Totals row
        totals_row = header_row + 1 + len(ALL_TIERS)
        totals_fill = PatternFill(
            start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid"
        )
        totals_font = Font(bold=True, color=_HEADER_FONT_COLOR, name=_FONT_SANS)
        for col_idx, value in enumerate(["TOTAL", total, "100%"], start=1):
            cell = ws.cell(row=totals_row, column=col_idx, value=value)
            cell.fill = totals_fill
            cell.font = totals_font

    # ------------------------------------------------------------------
    # Tab 3 — Executive Summary
    # ------------------------------------------------------------------

    def _build_exec_summary(
        self,
        wb: Workbook,
        vms: list[ClassifiedVM],
        customer_name: Optional[str],
    ) -> None:
        ws = wb.create_sheet("Executive Summary")

        # Logo row
        self._insert_logo(ws, row=1)

        total = len(vms)
        counts = Counter(vm.classification_tier for vm in vms)

        def pct(tier: str) -> str:
            return str(round(counts.get(tier, 0) / total * 100, 1)) if total > 0 else "0.0"

        customer_suffix = f" — {customer_name}" if customer_name else ""

        narrative = _EXEC_SUMMARY_TEMPLATE.format(
            customer_suffix=customer_suffix,
            date=datetime.now().strftime("%Y-%m-%d"),
            total=total,
            officially_supported_count=counts.get(TIER_OFFICIALLY_SUPPORTED, 0),
            officially_supported_pct=pct(TIER_OFFICIALLY_SUPPORTED),
            unofficially_supported_count=counts.get(TIER_UNOFFICIALLY_SUPPORTED, 0),
            unofficially_supported_pct=pct(TIER_UNOFFICIALLY_SUPPORTED),
            supported_vdi_count=counts.get(TIER_SUPPORTED_VDI, 0),
            supported_vdi_pct=pct(TIER_SUPPORTED_VDI),
            needs_review_count=counts.get(TIER_NEEDS_REVIEW, 0),
            needs_review_pct=pct(TIER_NEEDS_REVIEW),
            needs_info_count=counts.get(TIER_NEEDS_INFO, 0),
            needs_info_pct=pct(TIER_NEEDS_INFO),
            not_supported_count=counts.get(TIER_NOT_SUPPORTED, 0),
            not_supported_pct=pct(TIER_NOT_SUPPORTED),
        )

        start_row = 3
        for line_offset, line in enumerate(narrative.splitlines()):
            row_num = start_row + line_offset
            cell = ws.cell(row=row_num, column=1, value=line)
            cell.font = Font(name=_FONT_SANS, color="FFF9FAFB")

        ws.column_dimensions["A"].width = 90

    # ------------------------------------------------------------------
    # Logo helper
    # ------------------------------------------------------------------

    def _insert_logo(self, ws, row: int = 1) -> None:
        """Insert the US Signal logo into cell A{row}. Skips gracefully if missing."""
        if not _LOGO_PATH.exists():
            logger.warning(
                "Logo asset not found at %s — skipping logo insertion.", _LOGO_PATH
            )
            return

        try:
            img = XLImage(str(_LOGO_PATH))
            img.width = 120
            img.height = 40
            img.anchor = f"A{row}"
            ws.add_image(img)
            # Tint the logo cell background
            logo_fill = PatternFill(
                start_color=_LOGO_BG, end_color=_LOGO_BG, fill_type="solid"
            )
            ws.cell(row=row, column=1).fill = logo_fill
            ws.row_dimensions[row].height = 32
        except Exception:
            logger.warning("Failed to insert logo — skipping.", exc_info=True)
